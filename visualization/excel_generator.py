"""
Excel report generation with formatting.

Generates formatted Excel reports with multiple sheets, conditional formatting,
and professional styling.
"""

import pandas as pd
import numpy as np
from io import BytesIO
from typing import Dict, Any, List, Optional
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from config.settings import AppConfig

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generator for formatted Excel reports."""
    
    def __init__(self, config: Optional[AppConfig] = None):
        """
        Initialize Excel generator.
        
        Args:
            config: Application configuration
        """
        self.config = config or AppConfig()
    
    def generate(
        self,
        results: Dict[str, Any],
        input_data: pd.DataFrame
    ) -> BytesIO:
        """
        Generate comprehensive Excel report.
        
        Args:
            results: Optimization results for all scenarios
            input_data: Original input data
            
        Returns:
            BytesIO containing Excel file
        """
        output = BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write summary sheet
                self._write_summary_sheet(writer, results)
                
                # Write input data
                self._write_input_data_sheet(writer, input_data)
                
                # Write scenario sheets
                for scenario_name, result in results.items():
                    if result.get('success'):
                        self._write_scenario_sheets(
                            writer,
                            scenario_name,
                            result
                        )
                
                # Write comparison sheet if multiple scenarios
                if len([r for r in results.values() if r.get('success')]) >= 2:
                    self._write_comparison_sheet(writer, results)
            
            # Apply formatting
            output.seek(0)
            wb = load_workbook(output)
            self._apply_formatting(wb, results)
            
            # Save formatted workbook
            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)
            
            logger.info("Excel report generated successfully")
            return final_output
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}", exc_info=True)
            raise
    
    def _write_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any]
    ):
        """Write summary statistics sheet."""
        summary_data = []
        
        for scenario_name, result in results.items():
            if result.get('success'):
                summary_data.append({
                    'Scenario': scenario_name.replace('scenario_', 'Scenario '),
                    'Status': result.get('status', 'unknown'),
                    'Total Sample': result.get('total_sample', 0),
                    'Panel Sample': result.get('panel_sample', 0),
                    'Fresh Sample': result.get('fresh_sample', 0),
                    'Panel %': result.get('panel_pct', 0),
                    'Fresh %': result.get('fresh_pct', 0),
                    'Solver': result.get('solver', 'unknown'),
                    'Solve Time (s)': round(result.get('solve_time', 0), 2),
                    'Objective Value': round(result.get('objective_value', 0), 2)
                })
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    def _write_input_data_sheet(
        self,
        writer: pd.ExcelWriter,
        data: pd.DataFrame
    ):
        """Write original input data."""
        data.to_excel(writer, sheet_name='Input_Data', index=False)
    
    def _write_scenario_sheets(
        self,
        writer: pd.ExcelWriter,
        scenario_name: str,
        result: Dict[str, Any]
    ):
        """Write all sheets for a scenario."""
        prefix = scenario_name.replace('scenario_', 'S')
        
        # Detailed allocation
        if 'df_long' in result:
            sheet_name = f"{prefix}_Detail"
            result['df_long'].to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )
        
        # Panel allocation pivot
        if 'pivot_panel' in result:
            sheet_name = f"{prefix}_Panel"
            result['pivot_panel'].to_excel(writer, sheet_name=sheet_name)
        
        # Fresh allocation pivot
        if 'pivot_fresh' in result:
            sheet_name = f"{prefix}_Fresh"
            result['pivot_fresh'].to_excel(writer, sheet_name=sheet_name)
        
        # Combined samples and base weights
        if 'df_combined' in result:
            sheet_name = f"{prefix}_Combined"
            result['df_combined'].to_excel(writer, sheet_name=sheet_name)
        
        # Totals by dimension
        totals_data = []
        
        if 'region_totals' in result:
            totals_data.append(('Region', result['region_totals']))
        if 'size_totals' in result:
            totals_data.append(('Size', result['size_totals']))
        if 'industry_totals' in result:
            totals_data.append(('Industry', result['industry_totals']))
        
        if totals_data:
            sheet_name = f"{prefix}_Totals"
            row = 0
            for dim_name, df_totals in totals_data:
                # Write dimension header
                df_header = pd.DataFrame([[f"{dim_name} Totals"]])
                df_header.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=row,
                    index=False,
                    header=False
                )
                row += 1
                
                # Write totals
                df_totals.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=row,
                    index=False
                )
                row += len(df_totals) + 3
    
    def _write_comparison_sheet(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any]
    ):
        """Write comparison between scenarios."""
        # Get successful scenarios
        scenarios = [
            (name, result) for name, result in results.items()
            if result.get('success')
        ]
        
        if len(scenarios) < 2:
            return
        
        # Compare first two scenarios
        s1_name, s1_result = scenarios[0]
        s2_name, s2_result = scenarios[1]
        
        # Build comparison dataframe
        if 'df_combined' in s1_result and 'df_combined' in s2_result:
            df1 = s1_result['df_combined'].reset_index()
            df2 = s2_result['df_combined'].reset_index()
            
            # Find common columns
            common_cols = [c for c in df1.columns if c in df2.columns]
            
            # Calculate differences
            df_diff = df1[common_cols].copy()
            for col in common_cols:
                if pd.api.types.is_numeric_dtype(df1[col]):
                    df_diff[col] = df1[col] - df2[col]
            
            df_diff.to_excel(
                writer,
                sheet_name='Comparison',
                index=False
            )
    
    def _apply_formatting(
        self,
        wb,
        results: Dict[str, Any]
    ):
        """Apply formatting to workbook."""
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply conditional formatting to base weight columns
            self._apply_conditional_formatting(ws, sheet_name)
    
    def _apply_conditional_formatting(self, ws, sheet_name: str):
        """Apply conditional formatting to appropriate columns."""
        if 'Combined' in sheet_name or 'Comparison' in sheet_name:
            # Find BaseWeight columns
            header_row = [cell.value for cell in ws[1]]
            
            for idx, header in enumerate(header_row, 1):
                if header and 'BaseWeight' in str(header):
                    # Find data range
                    col_letter = get_column_letter(idx)
                    max_row = ws.max_row
                    
                    if max_row > 2:  # Has data beyond header
                        # Get values for color scale
                        values = []
                        for row in range(2, max_row + 1):
                            cell_value = ws[f"{col_letter}{row}"].value
                            if isinstance(cell_value, (int, float)):
                                values.append(cell_value)
                        
                        if values:
                            # Apply color scale
                            rule = ColorScaleRule(
                                start_type="num",
                                start_value=min(values),
                                start_color=self.config.EXCEL_COLOR_SCALE_LOW,
                                mid_type="percentile",
                                mid_value=50,
                                mid_color=self.config.EXCEL_COLOR_SCALE_MID,
                                end_type="num",
                                end_value=max(values),
                                end_color=self.config.EXCEL_COLOR_SCALE_HIGH
                            )
                            
                            ws.conditional_formatting.add(
                                f"{col_letter}2:{col_letter}{max_row}",
                                rule
                            )
