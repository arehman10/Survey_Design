import streamlit as st
import pandas as pd
import numpy as np
import cvxpy as cp
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import os
import uuid

###############################################################################
# CONFIGURATION & CONSTANTS
###############################################################################

SESSIONS_DIR = "sessions"

# Color scheme for inputs (blue) and outputs (black)
BLUE_INPUT = "0000FF"
YELLOW_HIGHLIGHT = "FFFF00"

###############################################################################
# SESSION MANAGEMENT
###############################################################################

def init_sessions_dir():
    """Ensure we have a 'sessions' subfolder to store JSON files."""
    if not os.path.exists(SESSIONS_DIR):
        os.makedirs(SESSIONS_DIR)

def save_session_to_file(session_id, data_dict):
    """Write data_dict to a JSON file in SESSIONS_DIR using session_id as filename."""
    init_sessions_dir()
    file_path = os.path.join(SESSIONS_DIR, f"{session_id}.json")
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data_dict, f, ensure_ascii=False, indent=2)

def load_session_from_file(session_id):
    """Read the JSON from SESSIONS_DIR. Return dict or None if not found."""
    file_path = os.path.join(SESSIONS_DIR, f"{session_id}.json")
    if not os.path.exists(file_path):
        return None
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)

###############################################################################
# CALCULATION HELPERS
###############################################################################

def compute_n_infinity(z_score, margin_of_error, p):
    """Calculate infinite population sample size."""
    return (z_score ** 2) * p * (1 - p) / (margin_of_error ** 2)

def compute_fpc_min(N, n_infinity):
    """Finite Population Correction-based recommended minimum sample for a group of size N."""
    if N <= 0:
        return 0
    return n_infinity / (1 + n_infinity / N)

###############################################################################
# OPTIMIZATION SOLVER
###############################################################################

def solve_optimization(
    cell_df,
    region_min_dict,
    size_min_dict,
    industry_min_dict,
    total_sample_target,
    panel_pct,
    fresh_pct,
    universe_df
):
    """
    Solve the optimization problem to allocate panel and fresh samples.
    
    Returns: dict with success flag and results
    """
    try:
        n_cells = len(cell_df)
        
        # Create decision variables
        panel = cp.Variable(n_cells, integer=True)
        fresh = cp.Variable(n_cells, integer=True)
        
        # Get base weights from universe
        base_weights = []
        for _, row in cell_df.iterrows():
            match = universe_df[
                (universe_df["Region"] == row["Region"]) &
                (universe_df["Size"] == row["Size"]) &
                (universe_df["Industry"] == row["Industry"])
            ]
            if len(match) > 0:
                base_weights.append(match.iloc[0]["BaseWeight"])
            else:
                base_weights.append(1.0)  # Default if not found
        
        base_weights = np.array(base_weights)
        
        # Objective: minimize weighted deviation from base weights
        total_sample = panel + fresh
        deviations = cp.multiply(base_weights, (total_sample - base_weights))
        objective = cp.Minimize(cp.sum_squares(deviations))
        
        # Constraints
        constraints = [
            panel >= 0,
            fresh >= 0,
            cp.sum(panel) == int(total_sample_target * panel_pct),
            cp.sum(fresh) == int(total_sample_target * fresh_pct)
        ]
        
        # Region minimums
        for region, min_val in region_min_dict.items():
            idx = cell_df[cell_df["Region"] == region].index
            if len(idx) > 0:
                constraints.append(cp.sum(total_sample[idx]) >= min_val)
        
        # Size minimums
        for size, min_val in size_min_dict.items():
            idx = cell_df[cell_df["Size"] == size].index
            if len(idx) > 0:
                constraints.append(cp.sum(total_sample[idx]) >= min_val)
        
        # Industry minimums
        for industry, min_val in industry_min_dict.items():
            idx = cell_df[cell_df["Industry"] == industry].index
            if len(idx) > 0:
                constraints.append(cp.sum(total_sample[idx]) >= min_val)
        
        # Solve
        problem = cp.Problem(objective, constraints)
        problem.solve(solver=cp.GLPK_MI, verbose=False)
        
        if problem.status not in ["optimal", "optimal_inaccurate"]:
            return {"success": False, "message": f"Optimization failed: {problem.status}"}
        
        # Extract results
        panel_vals = np.round(panel.value).astype(int)
        fresh_vals = np.round(fresh.value).astype(int)
        
        result_df = cell_df.copy()
        result_df["PanelSample"] = panel_vals
        result_df["FreshSample"] = fresh_vals
        result_df["SampleTotal"] = panel_vals + fresh_vals
        result_df["BaseWeight"] = base_weights
        
        return {
            "success": True,
            "df": result_df,
            "objective_value": problem.value
        }
        
    except Exception as e:
        return {"success": False, "message": f"Error: {str(e)}"}

###############################################################################
# EXCEL OUTPUT CREATION WITH INTERACTIVE FORMULAS
###############################################################################

def create_interactive_excel(
    scenario1_result,
    scenario2_result,
    config_data,
    filename="sample_allocation.xlsx"
):
    """
    Create an interactive Excel file where users can adjust parameters
    and see results update automatically using formulas.
    """
    
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        wb = writer.book
        
        # ===================================================================
        # SHEET 1: CONTROL PANEL (Interactive Parameters)
        # ===================================================================
        ws_control = wb.create_sheet("Control Panel", 0)
        
        # Header
        ws_control['A1'] = "SAMPLE ALLOCATION CONTROL PANEL"
        ws_control['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_control['A1'].fill = PatternFill(start_color="154360", end_color="154360", fill_type="solid")
        ws_control.merge_cells('A1:D1')
        
        # Configuration section
        row = 3
        ws_control[f'A{row}'] = "CONFIGURATION"
        ws_control[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Add parameters with formulas
        params = [
            ("Total Sample Target", config_data.get("total_sample", 1000), "Total number of samples to allocate"),
            ("Panel %", config_data.get("panel_pct", 0.7) * 100, "Percentage from panel (rest is fresh)"),
            ("Confidence Level %", config_data.get("confidence", 95), "Confidence level for sample size calculation"),
            ("Margin of Error %", config_data.get("margin", 5), "Margin of error for calculations"),
            ("Proportion (p)", config_data.get("proportion", 0.5), "Expected proportion for sample size"),
        ]
        
        for param_name, param_value, description in params:
            ws_control[f'A{row}'] = param_name
            ws_control[f'B{row}'] = param_value
            ws_control[f'B{row}'].font = Font(color=BLUE_INPUT, bold=True)
            ws_control[f'B{row}'].fill = PatternFill(start_color=YELLOW_HIGHLIGHT, end_color=YELLOW_HIGHLIGHT, fill_type="solid")
            ws_control[f'C{row}'] = description
            ws_control[f'C{row}'].font = Font(italic=True, size=9)
            row += 1
        
        # Calculated values
        row += 1
        ws_control[f'A{row}'] = "CALCULATED VALUES"
        ws_control[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Panel count formula
        ws_control[f'A{row}'] = "Panel Sample Count"
        ws_control[f'B{row}'] = f'=B4*(B5/100)'
        ws_control[f'B{row}'].number_format = '#,##0'
        row += 1
        
        # Fresh count formula
        ws_control[f'A{row}'] = "Fresh Sample Count"
        ws_control[f'B{row}'] = f'=B4-B{row-1}'
        ws_control[f'B{row}'].number_format = '#,##0'
        row += 1
        
        # Column widths
        ws_control.column_dimensions['A'].width = 25
        ws_control.column_dimensions['B'].width = 15
        ws_control.column_dimensions['C'].width = 50
        
        # ===================================================================
        # SHEET 2: SCENARIO COMPARISON
        # ===================================================================
        ws_compare = wb.create_sheet("Scenario Comparison")
        
        # Headers
        ws_compare['A1'] = "Comparison: Scenario 1 vs Scenario 2"
        ws_compare['A1'].font = Font(bold=True, size=14)
        ws_compare.merge_cells('A1:H1')
        
        # Column headers
        headers = ["Region", "Size", "Industry", "S1 Panel", "S1 Fresh", "S2 Panel", "S2 Fresh", "Difference"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_compare.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
        
        # Data
        if scenario1_result.get("success") and scenario2_result.get("success"):
            df1 = scenario1_result["df"]
            df2 = scenario2_result["df"]
            
            row = 4
            for idx in range(len(df1)):
                ws_compare.cell(row=row, column=1).value = df1.iloc[idx]["Region"]
                ws_compare.cell(row=row, column=2).value = df1.iloc[idx]["Size"]
                ws_compare.cell(row=row, column=3).value = df1.iloc[idx]["Industry"]
                ws_compare.cell(row=row, column=4).value = df1.iloc[idx]["PanelSample"]
                ws_compare.cell(row=row, column=5).value = df1.iloc[idx]["FreshSample"]
                ws_compare.cell(row=row, column=6).value = df2.iloc[idx]["PanelSample"]
                ws_compare.cell(row=row, column=7).value = df2.iloc[idx]["FreshSample"]
                # Difference formula
                ws_compare.cell(row=row, column=8).value = f'=(D{row}+E{row})-(F{row}+G{row})'
                row += 1
        
        # Auto-width columns
        for col in range(1, 9):
            ws_compare.column_dimensions[get_column_letter(col)].width = 15
        
        # ===================================================================
        # SHEET 3: SCENARIO 1 DETAILS
        # ===================================================================
        if scenario1_result.get("success"):
            ws_s1 = wb.create_sheet("Scenario 1 Details")
            df1 = scenario1_result["df"]
            
            # Write data with formulas
            ws_s1['A1'] = "Scenario 1: Detailed Allocation"
            ws_s1['A1'].font = Font(bold=True, size=14)
            ws_s1.merge_cells('A1:G1')
            
            # Headers
            headers = ["Region", "Size", "Industry", "Panel", "Fresh", "Total", "Base Weight"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_s1.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
            
            # Data rows
            for idx, row_data in df1.iterrows():
                row = idx + 4
                ws_s1.cell(row=row, column=1).value = row_data["Region"]
                ws_s1.cell(row=row, column=2).value = row_data["Size"]
                ws_s1.cell(row=row, column=3).value = row_data["Industry"]
                ws_s1.cell(row=row, column=4).value = row_data["PanelSample"]
                ws_s1.cell(row=row, column=5).value = row_data["FreshSample"]
                # Total formula
                ws_s1.cell(row=row, column=6).value = f'=D{row}+E{row}'
                ws_s1.cell(row=row, column=7).value = row_data["BaseWeight"]
                ws_s1.cell(row=row, column=7).number_format = '0.00'
            
            # Summary row with formulas
            summary_row = len(df1) + 5
            ws_s1.cell(row=summary_row, column=1).value = "TOTAL"
            ws_s1.cell(row=summary_row, column=1).font = Font(bold=True)
            ws_s1.cell(row=summary_row, column=4).value = f'=SUM(D4:D{summary_row-1})'
            ws_s1.cell(row=summary_row, column=5).value = f'=SUM(E4:E{summary_row-1})'
            ws_s1.cell(row=summary_row, column=6).value = f'=SUM(F4:F{summary_row-1})'
            
            for col in range(1, 8):
                ws_s1.column_dimensions[get_column_letter(col)].width = 15
        
        # ===================================================================
        # SHEET 4: SCENARIO 2 DETAILS
        # ===================================================================
        if scenario2_result.get("success"):
            ws_s2 = wb.create_sheet("Scenario 2 Details")
            df2 = scenario2_result["df"]
            
            # Write data with formulas (similar structure to Scenario 1)
            ws_s2['A1'] = "Scenario 2: Detailed Allocation"
            ws_s2['A1'].font = Font(bold=True, size=14)
            ws_s2.merge_cells('A1:G1')
            
            headers = ["Region", "Size", "Industry", "Panel", "Fresh", "Total", "Base Weight"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_s2.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
            
            for idx, row_data in df2.iterrows():
                row = idx + 4
                ws_s2.cell(row=row, column=1).value = row_data["Region"]
                ws_s2.cell(row=row, column=2).value = row_data["Size"]
                ws_s2.cell(row=row, column=3).value = row_data["Industry"]
                ws_s2.cell(row=row, column=4).value = row_data["PanelSample"]
                ws_s2.cell(row=row, column=5).value = row_data["FreshSample"]
                ws_s2.cell(row=row, column=6).value = f'=D{row}+E{row}'
                ws_s2.cell(row=row, column=7).value = row_data["BaseWeight"]
                ws_s2.cell(row=row, column=7).number_format = '0.00'
            
            summary_row = len(df2) + 5
            ws_s2.cell(row=summary_row, column=1).value = "TOTAL"
            ws_s2.cell(row=summary_row, column=1).font = Font(bold=True)
            ws_s2.cell(row=summary_row, column=4).value = f'=SUM(D4:D{summary_row-1})'
            ws_s2.cell(row=summary_row, column=5).value = f'=SUM(E4:E{summary_row-1})'
            ws_s2.cell(row=summary_row, column=6).value = f'=SUM(F4:F{summary_row-1})'
            
            for col in range(1, 8):
                ws_s2.column_dimensions[get_column_letter(col)].width = 15
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    output.seek(0)
    return output.getvalue()

###############################################################################
# MAIN STREAMLIT APP
###############################################################################

def main():
    st.set_page_config(page_title="Sample Allocation Optimizer", layout="wide")
    
    st.title("📊 Sample Allocation Optimizer")
    st.markdown("---")
    
    # Initialize session state
    if "session_id" not in st.session_state:
        st.session_state.session_id = str(uuid.uuid4())
    
    # File upload
    uploaded_file = st.file_uploader(
        "Upload Universe Excel File",
        type=["xlsx", "xls"],
        help="Upload the Excel file containing universe data"
    )
    
    if uploaded_file:
        # Read the universe data
        try:
            universe_df = pd.read_excel(uploaded_file)
            st.success(f"✅ Loaded {len(universe_df)} rows from universe file")
            
            # Display universe preview
            with st.expander("📋 Preview Universe Data"):
                st.dataframe(universe_df.head(20), use_container_width=True)
            
            # Configuration inputs
            st.markdown("### ⚙️ Configuration")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                total_sample = st.number_input("Total Sample Target", value=1000, min_value=1, step=10)
                panel_pct = st.slider("Panel %", 0.0, 1.0, 0.7, 0.05)
            
            with col2:
                confidence = st.number_input("Confidence Level %", value=95, min_value=1, max_value=99)
                margin = st.number_input("Margin of Error %", value=5, min_value=1, max_value=50)
            
            with col3:
                proportion = st.number_input("Proportion (p)", value=0.5, min_value=0.01, max_value=0.99, step=0.01)
            
            fresh_pct = 1.0 - panel_pct
            
            # Prepare cell data
            cell_df = universe_df[["Region", "Size", "Industry"]].drop_duplicates().reset_index(drop=True)
            
            # Calculate minimums (placeholder - you should adjust based on your logic)
            region_min_dict = {region: 50 for region in cell_df["Region"].unique()}
            size_min_dict = {size: 50 for size in cell_df["Size"].unique()}
            industry_min_dict = {industry: 50 for industry in cell_df["Industry"].unique()}
            
            # Run optimizations
            if st.button("🚀 Run Optimization", type="primary"):
                with st.spinner("Running Scenario 1..."):
                    scenario1 = solve_optimization(
                        cell_df, region_min_dict, size_min_dict, industry_min_dict,
                        total_sample, panel_pct, fresh_pct, universe_df
                    )
                
                with st.spinner("Running Scenario 2..."):
                    # Scenario 2 with slightly different parameters
                    scenario2 = solve_optimization(
                        cell_df, region_min_dict, size_min_dict, industry_min_dict,
                        total_sample, panel_pct * 0.9, fresh_pct * 1.1, universe_df
                    )
                
                if scenario1["success"] and scenario2["success"]:
                    st.success("✅ Optimization completed successfully!")
                    
                    # Display results
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("#### Scenario 1 Results")
                        st.dataframe(scenario1["df"], use_container_width=True)
                    
                    with col2:
                        st.markdown("#### Scenario 2 Results")
                        st.dataframe(scenario2["df"], use_container_width=True)
                    
                    # Generate Excel
                    config_data = {
                        "total_sample": total_sample,
                        "panel_pct": panel_pct,
                        "confidence": confidence,
                        "margin": margin,
                        "proportion": proportion
                    }
                    
                    excel_data = create_interactive_excel(
                        scenario1, scenario2, config_data
                    )
                    
                    st.download_button(
                        label="📥 Download Interactive Excel",
                        data=excel_data,
                        file_name="sample_allocation_interactive.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    if not scenario1["success"]:
                        st.error(f"Scenario 1 failed: {scenario1['message']}")
                    if not scenario2["success"]:
                        st.error(f"Scenario 2 failed: {scenario2['message']}")
        
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
    
    else:
        st.info("👆 Please upload an Excel file to begin")

if __name__ == "__main__":
    main()
